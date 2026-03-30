"""Handlers Telegram: lệnh, callback, hội thoại — logic DB nằm ở `services.py`."""

import re
import os
import html
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook

from AI.grader import grade_submission_with_ai
from AI.conduct_learning_evaluator import evaluate_conduct_and_learning_with_ai
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.error import BadRequest, Forbidden
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

from database import SessionLocal
from models import (
    Assignment,
    ClassInfo,
    StudentPoint,
    Submission,
    SubjectClass,
    SubjectInfo,
    UserInfo,
    UserRole,
)

from .services import (
    create_announcement_for_teacher,
    create_assignment_for_teacher,
    create_assignment_for_teacher_by_subject_class,
    create_join_request,
    get_assignment_submission_context,
    get_assignment_detail_for_student,
    get_teacher_assignment_detail,
    get_teacher_submission_detail,
    set_teacher_score_for_submission,
    get_or_create_user,
    list_classes_with_homeroom,
    list_active_assignments_for_student,
    list_homeroom_students_for_teacher,
    list_teacher_assignments_for_subject_class,
    list_teacher_submissions_for_assignment,
    list_pending_join_requests_for_homeroom_teacher,
    list_recent_announcements_for_student,
    list_students_in_class,
    list_teacher_subject_classes,
    list_teacher_managed_classes,
    list_student_submissions_results,
    list_teacher_classes_for_conduct,
    list_students_conduct_points_for_class,
    apply_student_conduct_points,
    resolve_join_request,
    student_join_block_reason,
    upsert_submission_for_student,
    upsert_submission_file_only_for_student,
    update_teacher_full_name,
    update_student_full_name,
    get_class_info,
    get_subject_info,
)

from .utils import normalize_stored_path, safe_path_component

# --- Nhãn nút (UI Telegram) ---
BTN_TEACHER_EDIT = "Chỉnh sửa thông tin"
BTN_TEACHER_ACCEPT_STUDENTS = "Chấp nhận học sinh"
BTN_TEACHER_ANNOUNCE = "Thông báo"
BTN_TEACHER_ASSIGN = "Giao bài"
BTN_TEACHER_HOMEROOM_STUDENTS = "Danh sách học sinh lớp chủ nhiệm"
BTN_TEACHER_MANAGED_CLASSES = "Các lớp quản lý"
BTN_TEACHER_GRADE = "Chấm bài"
BTN_TEACHER_CONDUCT = "Điểm rèn luyện học sinh"
BTN_TEACHER_ME = "Thông tin của tôi"
BTN_TEACHER_EXPORT_CONDUCT = "Xuất Excel quá trình rèn luyện học sinh lớp đang chủ nhiệm"

BTN_STUDENT_JOIN = "Tham gia lớp học"
BTN_STUDENT_EDIT = "Chỉnh sửa thông tin"
BTN_STUDENT_MENU = "Tính Năng"
BTN_STUDENT_ME = "Thông tin của tôi"
BTN_STUDENT_RECENT_ANNOUNCE = "Thông báo gần đây"
BTN_STUDENT_ASSIGNMENTS = "Bài tập"
BTN_STUDENT_RESULTS = "Xem kết quả bài tập"

CB_T_EDIT = "mt_edit"
CB_T_ACCEPT = "mt_accept"
CB_T_ANN = "mt_ann"
CB_T_ASSIGN = "mt_assign"
CB_T_HR = "mt_hr"
CB_T_CLS = "mt_cls"
CB_T_GRADE = "mt_grade"
CB_T_COND = "mt_cond"
CB_T_ME = "mt_me"
CB_T_XL = "mt_xl"

CB_S_JOIN = "ms_join"
CB_S_EDIT = "ms_edit"
CB_S_MENU = "ms_menu"
CB_S_ME = "ms_me"
CB_S_MENU_RECENT_ANNOUNCE = "ms_menu_ann"
CB_S_MENU_ASSIGNMENTS = "ms_menu_assign"
CB_S_MENU_RESULTS = "ms_menu_results"

# ms_jc_{class_id} — chọn lớp; mt_ar_{request_id}_a|r — duyệt/từ chối yêu cầu
RE_PICK_CLASS = re.compile(r"^ms_jc_(\d+)$")
RE_JOIN_RESOLVE = re.compile(r"^mt_ar_(\d+)_([ar])$")
RE_ANNOUNCE_CLASS_PICK = re.compile(r"^mt_ann_cls_(\d+)$")
RE_CONDUCT_CLASS_PICK = re.compile(r"^mt_cond_cls_(\d+)$")

WAITING_STUDENT_FULL_NAME = 1
WAITING_TEACHER_FULL_NAME = 2
WAITING_TEACHER_ANNOUNCE_TEXT = 3

PROMPT_STUDENT_FULL_NAME = (
    """Nhập tên đầy đủ của bạn:
    
(Lưu ý: giáo viên sẽ nhận biết bạn qua tên này.)
"""
)

PROMPT_TEACHER_FULL_NAME = (
    """Nhập tên giáo viên (họ tên đầy đủ) bạn muốn hiển thị trong hệ thống:"""
)

MENU_TEACHER_TEXT = (
    "📋 Menu giáo viên\n\n"
    "Chọn chức năng bên dưới. Các mục sẽ được nối logic trong các bước tiếp theo."
)
MENU_STUDENT_TEXT = "📋 Menu học sinh\n\nChọn chức năng bên dưới."


def teacher_menu_inline() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(BTN_TEACHER_EDIT, callback_data=CB_T_EDIT),
                InlineKeyboardButton(BTN_TEACHER_ACCEPT_STUDENTS, callback_data=CB_T_ACCEPT),
            ],
            [
                InlineKeyboardButton(BTN_TEACHER_ANNOUNCE, callback_data=CB_T_ANN),
                InlineKeyboardButton(BTN_TEACHER_ASSIGN, callback_data=CB_T_ASSIGN),
            ],
            [
                InlineKeyboardButton(BTN_TEACHER_ME, callback_data=CB_T_ME),
                InlineKeyboardButton(BTN_TEACHER_MANAGED_CLASSES, callback_data=CB_T_CLS),
            ],
            [
                InlineKeyboardButton(BTN_TEACHER_GRADE, callback_data=CB_T_GRADE),
                InlineKeyboardButton(BTN_TEACHER_CONDUCT, callback_data=CB_T_COND),
            ],
            [InlineKeyboardButton(BTN_TEACHER_HOMEROOM_STUDENTS, callback_data=CB_T_HR)],
            [InlineKeyboardButton(BTN_TEACHER_EXPORT_CONDUCT, callback_data=CB_T_XL)],
        ]
    )


def student_menu_inline() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton(BTN_STUDENT_JOIN, callback_data=CB_S_JOIN),
                InlineKeyboardButton(BTN_STUDENT_EDIT, callback_data=CB_S_EDIT),
            ],
            [
                InlineKeyboardButton(BTN_STUDENT_MENU, callback_data=CB_S_MENU),
                InlineKeyboardButton(BTN_STUDENT_ME, callback_data=CB_S_ME),
            ],
        ]
    )


async def _reply_or_edit_menu(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    *,
    text: str,
    reply_markup: InlineKeyboardMarkup,
) -> None:
    if update.callback_query:
        q = update.callback_query
        await q.answer()
        try:
            await q.edit_message_text(text, reply_markup=reply_markup)
        except BadRequest:
            if q.message:
                await context.bot.send_message(
                    chat_id=q.message.chat_id,
                    text=text,
                    reply_markup=reply_markup,
                )
        return
    if update.message:
        await update.message.reply_text(text, reply_markup=reply_markup)


async def send_role_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.effective_user:
        return
    if not update.message and not update.callback_query:
        return
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)

    if user.role == UserRole.TEACHER:
        await _reply_or_edit_menu(
            update,
            context,
            text=MENU_TEACHER_TEXT,
            reply_markup=teacher_menu_inline(),
        )
    else:
        await _reply_or_edit_menu(
            update,
            context,
            text=MENU_STUDENT_TEXT,
            reply_markup=student_menu_inline(),
        )


async def handle_student_join_classes(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Nút Tham gia lớp học — hiện lớp có GVCN."""
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.STUDENT:
            if q.message:
                await q.message.reply_text("Chức năng chỉ dành cho học sinh.")
            return
        reason = student_join_block_reason(db, user)
        if reason:
            if q.message:
                await q.message.reply_text(reason)
            return
        classes = list_classes_with_homeroom(db)
    if not classes:
        if q.message:
            await q.message.reply_text("Hiện chưa có lớp nào có giáo viên chủ nhiệm.")
        return
    kb = [[InlineKeyboardButton(c.name, callback_data=f"ms_jc_{c.id}")] for c in classes]
    if q.message:
        await q.message.reply_text(
            "Chọn lớp bạn muốn tham gia (đã có giáo viên chủ nhiệm):",
            reply_markup=InlineKeyboardMarkup(kb),
        )


async def on_student_pick_class(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Chọn lớp — tạo RequestJoinClass pending."""
    q = update.callback_query
    if not q or not q.data or not update.effective_user:
        return
    m = RE_PICK_CLASS.match(q.data)
    if not m:
        return
    class_id = int(m.group(1))
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        ok, msg = create_join_request(db, user, class_id)
    if q.message:
        await q.message.reply_text(msg)


async def handle_teacher_pending_join_list(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Nút Chấp nhận học sinh — danh sách yêu cầu + nút Chấp nhận / Từ chối."""
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            if q.message:
                await q.message.reply_text("Chức năng chỉ dành cho giáo viên.")
            return
        pending = list_pending_join_requests_for_homeroom_teacher(db, user.id)
    if not pending:
        if q.message:
            await q.message.reply_text(
                "Không có học sinh nào đang chờ duyệt vào các lớp bạn chủ nhiệm."
            )
        return
    lines = []
    for i, r in enumerate(pending, 1):
        lines.append(f"{i}. {r.student.full_name} → lớp {r.class_info.name}")
    text = "Học sinh đang yêu cầu tham gia lớp:\n\n" + "\n".join(lines)
    kb = []
    for r in pending:
        kb.append(
            [
                InlineKeyboardButton("Chấp nhận", callback_data=f"mt_ar_{r.id}_a"),
                InlineKeyboardButton("Từ chối", callback_data=f"mt_ar_{r.id}_r"),
            ]
        )
    if q.message:
        await q.message.reply_text(text, reply_markup=InlineKeyboardMarkup(kb))


async def handle_teacher_managed_classes(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            if q.message:
                await q.message.reply_text("Chức năng chỉ dành cho giáo viên.")
            return
        classes = list_teacher_managed_classes(db, user.id)
    if not classes:
        if q.message:
            await q.message.reply_text("Bạn chưa được phân công chủ nhiệm hoặc dạy lớp nào.")
        return
    lines = []
    for i, row in enumerate(classes, 1):
        roles = []
        if row["is_homeroom"]:
            roles.append("Chủ nhiệm")
        subjects = row["subjects"]
        if subjects:
            roles.append("Dạy: " + ", ".join(subjects))
        if not roles:
            roles.append("Đang quản lý")
        lines.append(
            f"{i}. {row['class_name']} (id: {row['class_id']})\n"
            f"   Vai trò: {' | '.join(roles)}"
        )
    text = "📚 Các lớp quản lý của bạn:\n\n" + "\n".join(lines)
    if q.message:
        await q.message.reply_text(text)


async def handle_teacher_homeroom_students(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            if q.message:
                await q.message.reply_text("Chức năng chỉ dành cho giáo viên.")
            return
        buckets = list_homeroom_students_for_teacher(db, user.id)
    if not buckets:
        if q.message:
            await q.message.reply_text("Bạn hiện không chủ nhiệm lớp nào.")
        return
    parts = []
    for bucket in buckets:
        parts.append(f"Lớp {bucket['class_name']} (id: {bucket['class_id']})")
        students = bucket["students"]
        if not students:
            parts.append("  (Chưa có học sinh)")
            parts.append("")
            continue
        for s in students:
            parts.append(
                f"{s.id} - {s.full_name} - {s.telegram_id} - {s.username or ''}"
            )
        parts.append("")
    text = "👥 Danh sách học sinh lớp chủ nhiệm:\n\n" + "\n".join(parts).rstrip()
    if q.message:
        await q.message.reply_text(text)


def _parse_assign_deadline(raw: str):
    for fmt in ("%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M"):
        try:
            dt = datetime.strptime(raw.strip(), fmt)
            return dt.replace(tzinfo=timezone(timedelta(hours=7))).astimezone(timezone.utc)
        except ValueError:
            continue
    return None


def _split_assign_payload(payload: str) -> tuple[bool, list[str] | str]:
    parts = [p.strip() for p in payload.split("|")]
    if len(parts) != 4:
        return (
            False,
            "Sai cú pháp. Dùng: /assign id_subject_class| tiêu đề| hướng dẫn| hạn chót",
        )
    return True, parts


async def handle_teacher_assign_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            if q.message:
                await q.message.reply_text("Chức năng chỉ dành cho giáo viên.")
            return
        rows = list_teacher_subject_classes(db, user.id)
    if not rows:
        if q.message:
            await q.message.reply_text("Bạn chưa được phân công dạy môn nào.")
        return
    introduce = "Các lớp và môn bạn đang dạy:\n\n"
    lines = [introduce + "idSubjectClass | Tên lớp | Tên môn"]
    for r in rows:
        lines.append(
            f"{r['subject_class_id']} | {r['class_name']} | {r['subject_name']}"
        )
    guide = (
        "\n\nCách giao bài:\n"
        "/assign id_subject_class| tiêu đề bài tập| hướng dẫn| hạn chót\n"
        "Ví dụ: /assign 10| Bài tập chương 1| Làm từ câu 1-5| 30/03/2026 23:00\n"
        "Bạn có thể đính kèm file .pdf (gửi cùng caption bắt đầu bằng /assign ...)."
    )
    if q.message:
        await q.message.reply_text("\n".join(lines) + guide)


async def teacher_assign_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    msg = update.message
    if not msg or not update.effective_user:
        return
    text = (msg.text or msg.caption or "").strip()
    if not text.startswith("/assign"):
        return
    payload = text[len("/assign"):].strip()
    ok_parse, data = _split_assign_payload(payload)
    if not ok_parse:
        await msg.reply_text(str(data))
        return
    subject_class_id_raw, title, instruction, deadline_raw = data

    try:
        subject_class_id = int(subject_class_id_raw)
    except ValueError:
        await msg.reply_text("id_subject_class phải là số nguyên.")
        return

    deadline = _parse_assign_deadline(deadline_raw)
    if deadline is None:
        await msg.reply_text(
            "Hạn chót không hợp lệ. Dùng dd/mm/yyyy HH:MM hoặc yyyy-mm-dd HH:MM."
        )
        return

    tg = update.effective_user
    ok_create = False
    resp = ""
    assignment_id: int | None = None

    teacher_name = ""
    class_id = 0
    class_name = ""
    subject_name = ""
    student_tele_ids: list[str] = []
    file_path: str | None = None

    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            await msg.reply_text("Chức năng chỉ dành cho giáo viên.")
            return

        teacher_name = user.full_name
        rows = list_teacher_subject_classes(db, user.id)
        picked = next(
            (r for r in rows if r["subject_class_id"] == subject_class_id),
            None,
        )
        if picked is None:
            await msg.reply_text("Bạn không được phân công dạy môn này.")
            return

        class_id = picked["class_id"]
        class_name = picked["class_name"]
        subject_name = picked["subject_name"]

        if msg.document:
            if msg.document.mime_type != "application/pdf":
                await msg.reply_text("Chỉ hỗ trợ file .pdf cho bài tập.")
                return
            upload_dir = os.path.join(
                "uploads",
                "assignments",
                str(class_name),
                str(subject_name),
            )
            os.makedirs(upload_dir, exist_ok=True)
            safe_name = f"{msg.document.file_unique_id}_{msg.document.file_name or 'assignment.pdf'}"
            local_path = os.path.join(upload_dir, safe_name)
            tg_file = await msg.document.get_file()
            await tg_file.download_to_drive(custom_path=local_path)
            file_path = local_path

        ok_create, resp, assignment_id = create_assignment_for_teacher_by_subject_class(
            db,
            teacher_user_id=user.id,
            subject_class_id=subject_class_id,
            title=title,
            instruction_text=instruction,
            deadline=deadline,
            file_path=file_path,
        )
        if ok_create and assignment_id is not None:
            student_tele_ids = [
                s.telegram_id for s in list_students_in_class(db, class_id)
            ]

    await msg.reply_text(resp)
    if not ok_create or assignment_id is None:
        return

    vn_tz = timezone(timedelta(hours=7))
    deadline_local = deadline.astimezone(vn_tz).strftime("%d/%m/%Y %H:%M")
    notify_text = (
        f"Giáo viên {html.escape(teacher_name)} vừa giao bài tập cho lớp {html.escape(class_name)}.\n"
        f"Môn: {html.escape(subject_name)}\n"
        f"Tiêu đề: {html.escape(title)}\n"
        f"Hạn chót: {html.escape(deadline_local)} (UTC+7)\n"
        f"<b>Hướng dẫn:</b>\n{html.escape(instruction)}\n\n"
        f"Nộp bài: gửi bài bằng file pdf, tên file là {assignment_id}.pdf"
    )
    doc_caption = (
        f"<b>📎 Đề bài tập</b> — tải file PDF về máy\n"
        f"• {html.escape(title)}\n"
        f"• Mã bài: {assignment_id}\n"
        "Chạm vào file ở trên để tải / lưu."
    )

    sent = 0
    for tele_id in student_tele_ids:
        try:
            await context.bot.send_message(
                chat_id=int(tele_id),
                text=notify_text,
                parse_mode="HTML",
            )
            if file_path:
                normalized_document_path = normalize_stored_path(file_path)
                with open(normalized_document_path, "rb") as f:
                    await context.bot.send_document(
                        chat_id=int(tele_id),
                        document=f,
                        caption=doc_caption,
                        parse_mode="HTML",
                    )
            sent += 1
        except (BadRequest, Forbidden):
            continue

    await msg.reply_text(
        f"Đã thông báo bài tập tới học sinh trong lớp ({sent}/{len(student_tele_ids)})."
    )


async def teacher_announce_start(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    q = update.callback_query
    if not q or not update.effective_user:
        return ConversationHandler.END
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            if q.message:
                await q.message.reply_text("Chức năng chỉ dành cho giáo viên.")
            return ConversationHandler.END
        classes = list_teacher_managed_classes(db, user.id)
    if not classes:
        if q.message:
            await q.message.reply_text("Bạn chưa có lớp để gửi thông báo.")
        return ConversationHandler.END
    kb = [
        [InlineKeyboardButton(c["class_name"], callback_data=f"mt_ann_cls_{c['class_id']}")]
        for c in classes
    ]
    if q.message:
        await q.message.reply_text(
            "Chọn lớp bạn muốn gửi thông báo:",
            reply_markup=InlineKeyboardMarkup(kb),
        )
    return WAITING_TEACHER_ANNOUNCE_TEXT


async def teacher_announce_pick_class(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    q = update.callback_query
    if not q or not q.data or not update.effective_user:
        return ConversationHandler.END
    m = RE_ANNOUNCE_CLASS_PICK.match(q.data)
    if not m:
        return WAITING_TEACHER_ANNOUNCE_TEXT
    class_id = int(m.group(1))
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        classes = list_teacher_managed_classes(db, user.id)
    picked = next((c for c in classes if c["class_id"] == class_id), None)
    if picked is None:
        if q.message:
            await q.message.reply_text("Bạn không có quyền gửi thông báo cho lớp này.")
        return ConversationHandler.END
    context.user_data["announce_class_id"] = class_id
    context.user_data["announce_class_name"] = picked["class_name"]
    if q.message:
        await q.message.reply_text(
            f"Hãy viết thông báo của bạn gửi tới lớp {picked['class_name']}"
        )
    return WAITING_TEACHER_ANNOUNCE_TEXT


async def teacher_announce_receive_text(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    if not update.message or not update.effective_user:
        return WAITING_TEACHER_ANNOUNCE_TEXT
    class_id = context.user_data.get("announce_class_id")
    class_name = context.user_data.get("announce_class_name")
    if not class_id or not class_name:
        await update.message.reply_text(
            "Bạn chưa chọn lớp. Hãy bấm lại nút Thông báo."
        )
        return ConversationHandler.END
    body = (update.message.text or "").strip()
    if not body:
        await update.message.reply_text("Vui lòng nhập nội dung thông báo.")
        return WAITING_TEACHER_ANNOUNCE_TEXT
    tg = update.effective_user
    with SessionLocal() as db:
        teacher = get_or_create_user(db, tg)
        ok, _ = create_announcement_for_teacher(db, teacher.id, class_id, body)
        if not ok:
            await update.message.reply_text(
                "Không thể tạo thông báo cho lớp này. Vui lòng thử lại."
            )
            context.user_data.pop("announce_class_id", None)
            context.user_data.pop("announce_class_name", None)
            return ConversationHandler.END
        teacher_name = teacher.full_name
        student_tele_ids = [s.telegram_id for s in list_students_in_class(db, class_id)]
    text = f"Giáo viên {teacher_name} thông báo: {body}"
    sent = 0
    for tele_id in student_tele_ids:
        try:
            await context.bot.send_message(chat_id=int(tele_id), text=text)
            sent += 1
        except (BadRequest, Forbidden):
            continue
    await update.message.reply_text(
        f"Đã gửi thông báo tới lớp {class_name} ({sent}/{len(student_tele_ids)} học sinh nhận được)."
    )
    context.user_data.pop("announce_class_id", None)
    context.user_data.pop("announce_class_name", None)
    return ConversationHandler.END


async def teacher_announce_cancel(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    context.user_data.pop("announce_class_id", None)
    context.user_data.pop("announce_class_name", None)
    if update.message:
        await update.message.reply_text("Đã hủy gửi thông báo.")
    return ConversationHandler.END


async def on_teacher_join_resolve(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Bấm Chấp nhận / Từ chối trên một yêu cầu — gửi thông báo cho học sinh."""
    q = update.callback_query
    if not q or not q.data or not update.effective_user:
        return
    m = RE_JOIN_RESOLVE.match(q.data)
    if not m:
        return
    request_id = int(m.group(1))
    approve = m.group(2) == "a"
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        ok, err_msg, student, cls = resolve_join_request(
            db, request_id, user.id, approve
        )
    if not ok:
        if q.message:
            await q.message.reply_text(err_msg)
        return

    class_name = cls.name if cls else ""
    if student:
        if approve:
            body = f"Bạn đã được giáo viên chủ nhiệm lớp {class_name} duyệt vào lớp."
        else:
            body = f"Bạn không được giáo viên chủ nhiệm lớp {class_name} duyệt vào lớp."
        try:
            await context.bot.send_message(chat_id=int(student.telegram_id), text=body)
        except (BadRequest, Forbidden):
            pass

    if q.message:
        await q.message.reply_text("Đã xử lý yêu cầu.")


async def student_edit_name_start(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    q = update.callback_query
    if not q or not update.effective_user:
        return ConversationHandler.END
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
    if user.role != UserRole.STUDENT:
        if q.message:
            await q.message.reply_text("Chức năng chỉ dành cho học sinh.")
        return ConversationHandler.END
    if q.message:
        await q.message.reply_text(PROMPT_STUDENT_FULL_NAME)
    return WAITING_STUDENT_FULL_NAME


async def student_edit_name_receive(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    if not update.message or not update.effective_user:
        return WAITING_STUDENT_FULL_NAME
    raw = (update.message.text or "").strip()
    if not raw:
        await update.message.reply_text("Vui lòng nhập tên đầy đủ.")
        return WAITING_STUDENT_FULL_NAME
    name = raw[:255]
    tg = update.effective_user
    with SessionLocal() as db:
        ok = update_student_full_name(db, tg, name)
    if not ok:
        await update.message.reply_text("Không tìm thấy tài khoản học sinh.")
        return ConversationHandler.END
    await update.message.reply_text(f"Đã xác nhận bạn là {name}")
    return ConversationHandler.END


async def student_edit_name_cancel(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    if update.message:
        await update.message.reply_text(
            "Đã hủy đổi tên. Dùng /start hoặc /menu nếu cần."
        )
    return ConversationHandler.END


async def teacher_edit_name_start(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    q = update.callback_query
    if not q or not update.effective_user:
        return ConversationHandler.END
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
    if user.role != UserRole.TEACHER:
        if q.message:
            await q.message.reply_text("Chức năng chỉ dành cho giáo viên.")
        return ConversationHandler.END
    if q.message:
        await q.message.reply_text(PROMPT_TEACHER_FULL_NAME)
    return WAITING_TEACHER_FULL_NAME


async def teacher_edit_name_receive(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    if not update.message or not update.effective_user:
        return WAITING_TEACHER_FULL_NAME
    raw = (update.message.text or "").strip()
    if not raw:
        await update.message.reply_text("Vui lòng nhập họ tên giáo viên.")
        return WAITING_TEACHER_FULL_NAME
    name = raw[:255]
    tg = update.effective_user
    with SessionLocal() as db:
        ok = update_teacher_full_name(db, tg, name)
    if not ok:
        await update.message.reply_text("Không tìm thấy tài khoản giáo viên.")
        return ConversationHandler.END
    await update.message.reply_text(f"Đã cập nhật tên giáo viên: {name}")
    return ConversationHandler.END


async def teacher_edit_name_cancel(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> int:
    if update.message:
        await update.message.reply_text(
            "Đã hủy đổi tên. Dùng /start hoặc /menu nếu cần."
        )
    return ConversationHandler.END


async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.effective_user or not update.message:
        return
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)

    if user.role == UserRole.TEACHER:
        await update.message.reply_text(
            "Xin chào thầy/cô! Tài khoản đã được đăng ký là giáo viên "
            "(telegram_id nằm trong whitelist TeleTeacherInfo).\n\n"
            f"Họ tên trong hệ thống: {user.full_name}",
            reply_markup=teacher_menu_inline(),
        )
    else:
        await update.message.reply_text(
            "Xin chào! Bạn đã được đăng ký là học sinh.\n\n"
            f"Họ tên trong hệ thống: {user.full_name}",
            reply_markup=student_menu_inline(),
        )


async def on_menu_button(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q or not q.data or not update.effective_user:
        return
    data = q.data
    tg = update.effective_user

    if data == CB_S_MENU:
        await handle_student_feature_menu(update, context)
        return

    if data == CB_S_JOIN:
        await handle_student_join_classes(update, context)
        return

    if data == CB_T_ACCEPT:
        await handle_teacher_pending_join_list(update, context)
        return

    if data == CB_T_ASSIGN:
        await handle_teacher_assign_menu(update, context)
        return

    if data == CB_T_GRADE:
        await handle_teacher_grade_menu(update, context)
        return

    if data == CB_T_COND:
        await handle_teacher_conduct_menu(update, context)
        return

    if data == CB_T_XL:
        await handle_teacher_export_conduct(update, context)
        return

    if data == CB_T_CLS:
        await handle_teacher_managed_classes(update, context)
        return

    if data == CB_T_HR:
        await handle_teacher_homeroom_students(update, context)
        return

    if data in (CB_T_ME, CB_S_ME):
        await handle_my_info(update, context)
        return

    if data == CB_S_MENU_RECENT_ANNOUNCE:
        await handle_student_recent_announcements(update, context)
        return

    if data == CB_S_MENU_ASSIGNMENTS:
        await handle_student_active_assignments(update, context)
        return

    if data == CB_S_MENU_RESULTS:
        await handle_student_results(update, context)
        return

    with SessionLocal() as db:
        user = get_or_create_user(db, tg)

    labels: dict[str, str] = {
        CB_T_EDIT: BTN_TEACHER_EDIT,
        CB_T_ANN: BTN_TEACHER_ANNOUNCE,
        CB_T_ASSIGN: BTN_TEACHER_ASSIGN,
        CB_T_HR: BTN_TEACHER_HOMEROOM_STUDENTS,
        CB_T_CLS: BTN_TEACHER_MANAGED_CLASSES,
        CB_T_GRADE: BTN_TEACHER_GRADE,
        CB_T_COND: BTN_TEACHER_CONDUCT,
        CB_T_ME: BTN_TEACHER_ME,
        CB_T_XL: BTN_TEACHER_EXPORT_CONDUCT,
        CB_S_ME: BTN_STUDENT_ME,
        CB_S_MENU_RECENT_ANNOUNCE: BTN_STUDENT_RECENT_ANNOUNCE,
        CB_S_MENU_ASSIGNMENTS: BTN_STUDENT_ASSIGNMENTS,
        CB_S_MENU_RESULTS: BTN_STUDENT_RESULTS,
    }

    if user.role == UserRole.TEACHER and data.startswith("ms_"):
        await q.answer("Chức năng dành cho học sinh.", show_alert=True)
        return
    if user.role == UserRole.STUDENT and data.startswith("mt_"):
        await q.answer("Chức năng dành cho giáo viên.", show_alert=True)
        return

    label = labels.get(data, data)
    await q.answer()
    if q.message:
        await context.bot.send_message(
            chat_id=q.message.chat_id,
            text=f"「{label}」 — tính năng sẽ được triển khai sau.",
        )


async def handle_student_recent_announcements(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.STUDENT:
            if q.message:
                await q.message.reply_text("Chức năng chỉ dành cho học sinh.")
            return
        if user.class_id is None:
            if q.message:
                await q.message.reply_text(
                    "Bạn chưa tham gia lớp nên chưa xem được thông báo."
                )
            return
        items = list_recent_announcements_for_student(db, user, days=7)
    if not items:
        if q.message:
            await q.message.reply_text(
                "Chưa có thông báo nào trong 7 ngày gần đây."
            )
        return
    blocks: list[str] = []
    for a in items:
        tname = a.teacher_info.full_name if a.teacher_info else "Giáo viên"
        ts = a.created_at
        if ts is not None and ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        dt_str = ts.strftime("%d/%m/%Y %H:%M") if ts else ""
        blocks.append(f"{dt_str} — {tname}\n{a.message}")
    text = "\n\n────────\n\n".join(blocks)
    header = "📢 Thông báo gần đây (7 ngày):\n\n"
    full = header + text
    max_len = 4096
    if q.message:
        for i in range(0, len(full), max_len):
            await q.message.reply_text(full[i : i + max_len])


async def handle_student_feature_menu(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    kb = [
        [
            InlineKeyboardButton(
                BTN_STUDENT_RECENT_ANNOUNCE, callback_data=CB_S_MENU_RECENT_ANNOUNCE
            ),
            InlineKeyboardButton(
                BTN_STUDENT_ASSIGNMENTS, callback_data=CB_S_MENU_ASSIGNMENTS
            ),
        ],
        [
            InlineKeyboardButton(
                BTN_STUDENT_RESULTS, callback_data=CB_S_MENU_RESULTS
            )
        ],
    ]
    if q.message:
        await q.message.reply_text(
            "📂 Tính năng cho học sinh\n\nChọn một mục bên dưới:",
            reply_markup=InlineKeyboardMarkup(kb),
        )


async def handle_student_active_assignments(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.STUDENT:
            if q.message:
                await q.message.reply_text("Chức năng chỉ dành cho học sinh.")
            return
        items = list_active_assignments_for_student(db, user)
    if not items:
        if q.message:
            await q.message.reply_text("Hiện không có bài tập còn thời hạn.")
        return
    vn_tz = timezone(timedelta(hours=7))
    lines = ["id bài tập | tên môn học | tiêu đề | hạn chót (UTC+7)"]
    for it in items:
        dl = it["deadline"]
        if dl is not None and dl.tzinfo is None:
            dl = dl.replace(tzinfo=timezone.utc)
        dl_str = dl.astimezone(vn_tz).strftime("%d/%m/%Y %H:%M") if dl else ""
        lines.append(f"{it['id']} | {it['subject_name']} | {it['title']} | {dl_str}")
    lines.append("\nXem chi tiết: /assignment {id bài tập}")
    if q.message:
        await q.message.reply_text("\n".join(lines))


async def handle_student_results(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.STUDENT:
            if q.message:
                await q.message.reply_text("Chức năng chỉ dành cho học sinh.")
            return
        items = list_student_submissions_results(db, user)
    if not items:
        if q.message:
            await q.message.reply_text("Bạn chưa có kết quả bài tập nào.")
        return

    vn_tz = timezone(timedelta(hours=7))
    blocks: list[str] = []
    for it in items:
        submitted_at = it.get("submitted_at")
        if submitted_at is not None and submitted_at.tzinfo is None:
            submitted_at = submitted_at.replace(tzinfo=timezone.utc)
        submitted_str = submitted_at.astimezone(vn_tz).strftime("%d/%m/%Y %H:%M") if submitted_at else ""

        teacher_score = it.get("teacher_score")
        ai_score = it.get("ai_score")
        teacher_txt = f"{teacher_score}/10" if teacher_score is not None else "Chưa chấm"
        ai_txt = f"{ai_score}/10" if ai_score is not None else "Chưa có AI chấm"

        blocks.append(
            "📌 "
            f"Môn: {it['subject_name']}\n"
            f"Tiêu đề: {it['assignment_title']}\n"
            f"Điểm giáo viên: {teacher_txt}\n"
            f"Điểm AI: {ai_txt}\n"
            f"Thời gian nộp: {submitted_str}\n\n"
        )

    # Cắt theo giới hạn 4096 ký tự
    max_len = 4096
    header = "📊 Kết quả học tập:\n"
    full = header + "\n".join(blocks)
    if q.message:
        for i in range(0, len(full), max_len):
            await q.message.reply_text(full[i : i + max_len])


async def student_assignment_detail_command(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    msg = update.message
    if not msg or not update.effective_user:
        return
    if not context.args:
        await msg.reply_text("Dùng cú pháp: /assignment {id bài tập}")
        return
    try:
        assignment_id = int(context.args[0].strip())
    except ValueError:
        await msg.reply_text("id bài tập phải là số nguyên.")
        return
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.STUDENT:
            await msg.reply_text("Chức năng chỉ dành cho học sinh.")
            return
        assignment, subject_name = get_assignment_detail_for_student(db, user, assignment_id)
    if assignment is None:
        await msg.reply_text("Không tìm thấy bài tập hoặc bạn không thuộc lớp của bài này.")
        return

    vn_tz = timezone(timedelta(hours=7))
    dl = assignment.deadline
    if dl is not None and dl.tzinfo is None:
        dl = dl.replace(tzinfo=timezone.utc)
    dl_str = dl.astimezone(vn_tz).strftime("%d/%m/%Y %H:%M") if dl else "Không có"
    text = (
        f"📝 Chi tiết bài tập\n\n"
        f"id bài tập: {assignment.id}\n"
        f"môn học: {subject_name or ''}\n"
        f"tiêu đề: {assignment.title}\n"
        f"hướng dẫn: \n{assignment.instruction_text}\n"
        f"hạn chót: {dl_str} (UTC+7)\n\n"
        f"Hướng dẫn nộp bài:\n"
        f"- Chuẩn bị file PDF bài làm.\n"
        f"- Đặt tên file: {assignment.id}.pdf\n"
        f"- Gửi file cho bot để nộp bài."
    )
    await msg.reply_text(text)
    if assignment.file_path:
        normalized = normalize_stored_path(assignment.file_path)
        try:
            with open(normalized, "rb") as f:
                await msg.reply_document(
                    document=f,
                    caption=f"📎 File bài tập id {assignment.id}",
                )
        except OSError:
            await msg.reply_text("Không đọc được file bài tập đính kèm.")


async def student_submit_assignment_pdf(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    msg = update.message
    if not msg or not update.effective_user or not msg.document:
        return
    if msg.caption and msg.caption.strip().startswith("/assign"):
        return

    doc = msg.document
    if (doc.file_name or "").lower().endswith(".pdf") is False:
        return
    if (doc.file_size or 0) > 10 * 1024 * 1024:
        await msg.reply_text("File nộp vượt quá 10MB. Em vui lòng nén hoặc tách file.")
        return

    m = re.match(r"^(\d+)\.pdf$", (doc.file_name or "").strip())
    if not m:
        await msg.reply_text(
            "Tên file nộp chưa đúng. Em hãy đặt tên theo dạng {id bài tập}.pdf, ví dụ: 15.pdf."
        )
        return
    assignment_id = int(m.group(1))

    tg = update.effective_user
    student_id = 0
    class_name = ""
    subject_name = ""
    assignment_title = ""
    assignment_instruction = ""
    assignment_file_path = None
    local_path = ""
    with SessionLocal() as db:
        student = get_or_create_user(db, tg)
        if student.role != UserRole.STUDENT:
            return
        student_id = student.id
        ctx = get_assignment_submission_context(db, student, assignment_id)
        if ctx is None:
            await msg.reply_text("Em không có quyền nộp bài này hoặc bài không tồn tại.")
            return
        deadline = ctx["deadline"]
        if deadline is not None:
            dl = deadline
            if dl.tzinfo is None:
                dl = dl.replace(tzinfo=timezone.utc)
            if dl < datetime.now(timezone.utc):
                await msg.reply_text("Bài tập này đã quá hạn nộp.")
                return
        class_name = safe_path_component(str(ctx["class_name"]))
        subject_name = safe_path_component(str(ctx["subject_name"]))
        student_name = safe_path_component(student.full_name)
        assignment_title = str(ctx["assignment_title"])
        assignment_instruction = str(ctx["assignment_instruction"])
        assignment_file_path = ctx["assignment_file_path"]
        submit_dir = os.path.join("uploads", "submissions", class_name, subject_name)
        os.makedirs(submit_dir, exist_ok=True)
        local_path = os.path.join(
            submit_dir,
            f"{assignment_id}_{student.id}_{student_name}.pdf",
        )

    tg_file = await doc.get_file()
    await tg_file.download_to_drive(custom_path=local_path)

    feedback, score = grade_submission_with_ai(
        subject_name=subject_name,
        assignment_title=assignment_title,
        assignment_instruction=assignment_instruction,
        assignment_file_path=assignment_file_path,
        submission_file_path=local_path,
    )

    if feedback == "__AI_READ_FAILED__":
        with SessionLocal() as db:
            ok, result_msg = upsert_submission_file_only_for_student(
                db,
                assignment_id=assignment_id,
                student_id=student_id,
                file_path=local_path,
            )
        if not ok:
            await msg.reply_text("Không thể lưu bài nộp. Em vui lòng thử lại.")
            return
        await msg.reply_text(
            "Chúc mừng em vì đã nộp bài đúng hạn và phần chấm bài sẽ phải đợi giáo viên chủ nhiệm chấm."
        )
        return

    with SessionLocal() as db:
        ok, result_msg = upsert_submission_for_student(
            db,
            assignment_id=assignment_id,
            student_id=student_id,
            file_path=local_path,
            ai_feedback=feedback,
            ai_score=score,
        )

    if not ok:
        await msg.reply_text("Không thể lưu bài nộp. Em vui lòng thử lại.")
        return

    score_txt = f"\nĐiểm tham khảo (AI): {score:.1f}/10" if score is not None else ""
    await msg.reply_text(
        f"{result_msg}\n\nNhận xét AI:\n{feedback}{score_txt}"
    )


def _utc7(dt: datetime | None) -> str:
    if dt is None:
        return "Không có"
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone(timedelta(hours=7))).strftime("%d/%m/%Y %H:%M")


async def handle_teacher_grade_menu(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            await q.message.reply_text("Chức năng chỉ dành cho giáo viên.")  # type: ignore[union-attr]
            return
        rows = list_teacher_subject_classes(db, user.id)

    if not rows:
        if q.message:
            await q.message.reply_text("Bạn chưa có lớp/môn nào để chấm bài.")
        return

    lines = ["idSubjectClass | Tên lớp | Tên môn"]
    for r in rows:
        lines.append(f"{r['subject_class_id']} | {r['class_name']} | {r['subject_name']}")
    lines.append("\nĐể xem danh sách bài tập đã giao của môn/lớp này, dùng lệnh:")
    lines.append("/viewAssigments {idSubjectClass}")
    if q.message:
        await q.message.reply_text("\n".join(lines))


async def teacher_view_assignments_command(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    msg = update.message
    if not msg or not update.effective_user:
        return
    if not context.args:
        await msg.reply_text("Dùng cú pháp: /viewAssigments {idSubjectClass}")
        return
    try:
        subject_class_id = int(context.args[0].strip())
    except ValueError:
        await msg.reply_text("idSubjectClass phải là số nguyên.")
        return
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            await msg.reply_text("Chức năng chỉ dành cho giáo viên.")
            return
        items = list_teacher_assignments_for_subject_class(db, user.id, subject_class_id)

    if not items:
        await msg.reply_text("Không có bài tập nào được giao cho môn/lớp này.")
        return

    lines = ["idAssignment | Tiêu đề | Hạn chót (UTC+7)"]
    for a in items:
        lines.append(f"{a['id']} | {a['title']} | {_utc7(a.get('deadline'))}")
    
    lines.append("\nĐể xem chi tiết bài tập cụ thể, dùng lệnh:")
    lines.append("/detailAssignment {idAssignment}")
    lines.append("\nĐể xem danh sách học sinh nộp bài cụ thể, dùng lệnh:")
    lines.append("/viewSubmissionsAssginment {idAssignment}")
    await msg.reply_text("\n".join(lines))


async def teacher_detail_assignment_command(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    msg = update.message
    if not msg or not update.effective_user:
        return
    if not context.args:
        await msg.reply_text("Dùng cú pháp: /detailAssignment {idAssignment}")
        return
    try:
        assignment_id = int(context.args[0].strip())
    except ValueError:
        await msg.reply_text("idAssignment phải là số nguyên.")
        return
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            await msg.reply_text("Chức năng chỉ dành cho giáo viên.")
            return
        detail = get_teacher_assignment_detail(db, user.id, assignment_id)
    if detail is None:
        await msg.reply_text("Không tìm thấy bài tập hoặc bạn không có quyền xem.")
        return
    assignment = detail["assignment"]
    file_path = assignment.file_path

    text = (
        "📝 Chi tiết assignment\n\n"
        f"id: {assignment.id}\n"
        f"môn: {detail['subject_name']}\n"
        f"lớp: {detail['class_name']}\n"
        f"tiêu đề: {assignment.title}\n"
        f"hướng dẫn: {assignment.instruction_text}\n"
        f"ngày giao: {_utc7(assignment.created_at)}\n"
        f"hạn chót: {_utc7(assignment.deadline)}\n\n"
        "Hướng dẫn nộp bài:\n"
        f"- Tên file nộp: {assignment.id}.pdf\n"
        "- Nộp PDF cho bot trong khung chat."
    )
    await msg.reply_text(text)
    if file_path:
        normalized = normalize_stored_path(file_path)
        try:
            with open(normalized, "rb") as f:
                await msg.reply_document(
                    document=f,
                    caption=f"📎 File đề bài tập (id {assignment.id})",
                )
        except OSError:
            await msg.reply_text("Không đọc được file đề bài (file_path không hợp lệ).")


async def teacher_view_submissions_assignment_command(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    msg = update.message
    if not msg or not update.effective_user:
        return
    if not context.args:
        await msg.reply_text(
            "Dùng cú pháp: /viewSubmissionsAssginment {idAssignment}"
        )
        return
    try:
        assignment_id = int(context.args[0].strip())
    except ValueError:
        await msg.reply_text("idAssignment phải là số nguyên.")
        return
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            await msg.reply_text("Chức năng chỉ dành cho giáo viên.")
            return
        items = list_teacher_submissions_for_assignment(db, user.id, assignment_id)

    if not items:
        await msg.reply_text("Chưa có học sinh nào nộp bài cho assignment này.")
        return

    lines = ["idSubmission | Họ tên học sinh"]
    for it in items:
        lines.append(f"{it['id']} | {it['student_name']}")
    # await msg.reply_text("\n".join(lines))    
    lines.append("\n\nĐể xem chi tiết bài nộp cụ thể của một học sinh, dùng lệnh:")
    lines.append("/viewSubmisson {idSubmission}")
    await msg.reply_text("\n".join(lines))


async def teacher_view_submission_command(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    msg = update.message
    if not msg or not update.effective_user:
        return
    if not context.args:
        await msg.reply_text("Dùng cú pháp: /viewSubmisson {idSubmission}")
        return
    try:
        submission_id = int(context.args[0].strip())
    except ValueError:
        await msg.reply_text("idSubmission phải là số nguyên.")
        return
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            await msg.reply_text("Chức năng chỉ dành cho giáo viên.")
            return
        detail = get_teacher_submission_detail(db, user.id, submission_id)
    if detail is None:
        await msg.reply_text("Không tìm thấy submission hoặc bạn không có quyền xem.")
        return
    sub = detail["submission"]
    assignment = detail["assignment"]

    text = (
        "📄 Chi tiết submission\n\n"
        f"idSubmission: {sub.id}\n"
        f"idAssignment: {sub.assignment_id}\n"
        f"tiêu đề: {assignment.title}\n"
        f"họ tên học sinh: {detail['student_name']}\n"
        f"file đã nộp: {sub.file_path}\n"
        f"thời gian nộp: {_utc7(sub.submitted_at)}\n\n"
        "AI feedback:\n"
        f"{sub.ai_feedback or ''}\n"
    )
    text += f"Điểm giáo viên: {sub.teacher_score or 'Không có'}"
    text += "\n\nĐể chấm điểm, dùng lệnh: /score {idSubmission} {điểm} (0-10)"
    await msg.reply_text(text)
    if sub.file_path:
        normalized = normalize_stored_path(sub.file_path)
        try:
            with open(normalized, "rb") as f:
                await msg.reply_document(
                    document=f,
                    caption=f"📎 File nộp (submission id {sub.id})",
                )
        except OSError:
            await msg.reply_text("Không đọc được file nộp (file_path không hợp lệ).")


async def teacher_score_submission_command(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    msg = update.message
    if not msg or not update.effective_user:
        return
    if len(context.args) < 2:
        await msg.reply_text("Dùng cú pháp: /score {id submission} {điểm} (0-10)")
        return
    try:
        submission_id = int(context.args[0].strip())
        score = float(context.args[1].strip())
    except ValueError:
        await msg.reply_text("id submission phải là số nguyên và điểm phải là số.")
        return

    if score < 0 or score > 10:
        await msg.reply_text("Điểm phải nằm trong khoảng từ 0 đến 10.")
        return

    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            await msg.reply_text("Chức năng chỉ dành cho giáo viên.")
            return
        ok, resp = set_teacher_score_for_submission(
            db, teacher_user_id=user.id, submission_id=submission_id, score=score
        )
    if ok:
        await msg.reply_text(resp)
    else:
        await msg.reply_text(resp)


async def handle_teacher_conduct_menu(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            await q.message.reply_text("Chức năng chỉ dành cho giáo viên.")  # type: ignore[union-attr]
            return
        classes = list_teacher_classes_for_conduct(db, user.id)
    if not classes:
        if q.message:
            await q.message.reply_text("Bạn chưa được phân công dạy hoặc chủ nhiệm lớp nào.")
        return
    kb = [
        [
            InlineKeyboardButton(
                c["class_name"], callback_data=f"mt_cond_cls_{c['class_id']}"
            )
        ]
        for c in classes
    ]
    if q.message:
        await q.message.reply_text(
            "Chọn lớp để quản lý điểm rèn luyện:",
            reply_markup=InlineKeyboardMarkup(kb),
        )


async def on_teacher_pick_conduct_class(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    q = update.callback_query
    if not q or not q.data or not update.effective_user:
        return
    m = RE_CONDUCT_CLASS_PICK.match(q.data)
    if not m:
        return
    class_id = int(m.group(1))
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            if q.message:
                await q.message.reply_text("Chức năng chỉ dành cho giáo viên.")
            return
        # Lấy tên lớp
        cls_row = db.query(ClassInfo.name).filter(ClassInfo.id == class_id).first()
        class_name = cls_row[0] if cls_row else str(class_id)
        students = list_students_conduct_points_for_class(db, user.id, class_id)
    if not students:
        if q.message:
            await q.message.reply_text("Không có học sinh hoặc bạn không có quyền xem lớp này.")
        return
    lines = [f"👥 Học sinh lớp {class_name}\n"]
    for s in students:
        lines.append(
            f"{s['student_id']} - {s['full_name']} - {s['telegram_id']} - {s['username'] or ''}"
        )
    lines.append("")
    lines.append("Cộng/trừ điểm rèn luyện:")
    lines.append("/plus {id học sinh} {lý do}")
    lines.append("Ví dụ: /plus 1234567890 Chăm phát biểu")
    lines.append("/minus {id học sinh} {lý do}")
    lines.append("Ví dụ: /minus 1234567890 Vắng học")
    if q.message:
        await q.message.reply_text("\n".join(lines))


async def teacher_plus_command(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    msg = update.message
    if not msg or not update.effective_user:
        return
    raw = (msg.text or "").strip()
    if not raw.startswith("/plus"):
        return
    after = raw[len("/plus") :].strip()
    parts = after.split(" ", 1)
    if len(parts) < 2:
        await msg.reply_text("Dùng cú pháp: /plus {id học sinh} {lý do}")
        return
    id_str, reason = parts[0].strip(), parts[1].strip()
    id_str = id_str.strip()
    reason = reason.strip()
    if not id_str or not reason:
        await msg.reply_text("Dùng cú pháp: /plus {id học sinh} {lý do}")
        return
    try:
        student_id = int(id_str)
    except ValueError:
        await msg.reply_text("id học sinh phải là số nguyên.")
        return
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            await msg.reply_text("Chức năng chỉ dành cho giáo viên.")
            return
        ok, resp, student_tele_id = apply_student_conduct_points(
            db,
            teacher_user_id=user.id,
            student_id=student_id,
            delta=1,
            reason=reason,
        )
    if not ok or student_tele_id is None:
        await msg.reply_text(resp)
        return
    await msg.reply_text(resp)
    try:
        await context.bot.send_message(
            chat_id=int(student_tele_id),
            text=f"Bạn được cộng điểm rèn luyện với lý do: {reason}",
        )
    except (BadRequest, Forbidden):
        pass


async def teacher_minus_command(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    msg = update.message
    if not msg or not update.effective_user:
        return
    raw = (msg.text or "").strip()
    if not raw.startswith("/minus"):
        return
    after = raw[len("/minus") :].strip()
    parts = after.split(" ", 1)
    if len(parts) < 2:
        await msg.reply_text("Dùng cú pháp: /minus {id học sinh} {lý do}")
        return
    id_str, reason = parts[0].strip(), parts[1].strip()
    if not id_str or not reason:
        await msg.reply_text("Dùng cú pháp: /minus {id học sinh} {lý do}")
        return
    try:
        student_id = int(id_str)
    except ValueError:
        await msg.reply_text("id học sinh phải là số nguyên.")
        return
    tg = update.effective_user
    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            await msg.reply_text("Chức năng chỉ dành cho giáo viên.")
            return
        ok, resp, student_tele_id = apply_student_conduct_points(
            db,
            teacher_user_id=user.id,
            student_id=student_id,
            delta=-1,
            reason=reason,
        )
    if not ok or student_tele_id is None:
        await msg.reply_text(resp)
        return
    await msg.reply_text(resp)
    try:
        await context.bot.send_message(
            chat_id=int(student_tele_id),
            text=f"Bạn bị nhắn nhở với lý do: {reason}",
        )
    except (BadRequest, Forbidden):
        pass


def _excel_safe_sheet_title(title: str) -> str:
    """Excel sheet title không được chứa: : \\ / ? * [ ]."""
    bad = [":", "\\", "/", "?", "*", "[", "]"]
    out = title or ""
    for ch in bad:
        out = out.replace(ch, "")
    out = out.strip()
    if not out:
        out = "Sheet"
    return out[:31]


def _conduct_comment(cong: int, nhac: int) -> str:
    if cong <= 0 and nhac <= 0:
        return "Chưa có dữ liệu rèn luyện."
    if nhac == 0 and cong > 0:
        return "Tích cực rèn luyện, chấp hành tốt."
    if cong > nhac:
        return "Rèn luyện tích cực, tuy nhiên vẫn còn một vài lần nhắc nhở."
    if cong == nhac and cong > 0:
        return "Có tiến bộ, cần tiếp tục duy trì và giảm vi phạm."
    return "Chưa tích cực rèn luyện; cần cố gắng hơn để giảm nhắc nhở."


def _learning_comment(best_subject: str | None, overall_avg: float | None) -> str:
    if not best_subject:
        return "Chưa đủ dữ liệu để nhận xét thế mạnh."
    if overall_avg is None:
        return f"Thế mạnh ở môn {best_subject}. Em cố gắng duy trì nhé."
    if overall_avg >= 8.0:
        return f"Thế mạnh ở môn {best_subject}. Kết quả học tập rất tốt."
    if overall_avg >= 6.5:
        return f"Thế mạnh ở môn {best_subject}. Em cần tiếp tục cố gắng để tiến bộ."
    return f"Thế mạnh ở môn {best_subject}. Em nên tập trung hơn để nâng điểm."


async def handle_teacher_export_conduct(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    tg = update.effective_user

    # Variables sẽ dùng tiếp sau khi đóng session
    students = []
    student_ids: list[int] = []
    assignment_by_subject_name: dict[str, list] = {}
    sub_map: dict[tuple[int, int], Submission] = {}
    points_map: dict[int, tuple[int, int]] = {}
    class_id = 0
    class_name = ""

    with SessionLocal() as db:
        user = get_or_create_user(db, tg)
        if user.role != UserRole.TEACHER:
            await q.message.reply_text("Tính năng chỉ dành cho giáo viên chủ nhiệm.")
            return

        homeroom_classes = (
            db.query(ClassInfo.id, ClassInfo.name)
            .filter(ClassInfo.homeroom_teacher_id == user.id)
            .order_by(ClassInfo.name.asc())
            .all()
        )
        if not homeroom_classes:
            await q.message.reply_text("tính năng chỉ dành cho giáo viên chủ nhiệm")
            return

        class_id, class_name = homeroom_classes[0]

        students = (
            db.query(UserInfo.id, UserInfo.full_name)
            .filter(UserInfo.class_id == class_id, UserInfo.role == UserRole.STUDENT)
            .order_by(UserInfo.full_name.asc())
            .all()
        )
        student_ids = [s.id for s in students]

        subject_rows = (
            db.query(SubjectClass.id, SubjectInfo.name)
            .join(SubjectInfo, SubjectClass.subject_id == SubjectInfo.id)
            .filter(SubjectClass.class_id == class_id)
            .order_by(SubjectInfo.name.asc(), SubjectClass.id.asc())
            .all()
        )
        subject_class_ids = [r.id for r in subject_rows]
        subject_map = {r.id: r.name for r in subject_rows}

        assignment_rows = (
            db.query(Assignment.id, Assignment.title, Assignment.subject_class_id)
            .filter(Assignment.subject_class_id.in_(subject_class_ids))
            .order_by(Assignment.subject_class_id.asc(), Assignment.id.asc())
            .all()
        )
        assignment_ids = [r.id for r in assignment_rows]

        for a in assignment_rows:
            subj_name = subject_map.get(a.subject_class_id, "Môn")
            assignment_by_subject_name.setdefault(subj_name, []).append(a)

        if assignment_ids and student_ids:
            subs = (
                db.query(Submission)
                .filter(Submission.assignment_id.in_(assignment_ids))
                .filter(Submission.student_id.in_(student_ids))
                .all()
            )
            sub_map = {(s.assignment_id, s.student_id): s for s in subs}

        # StudentPoint
        points_map = {sid: (0, 0) for sid in student_ids}
        if student_ids:
            points = db.query(StudentPoint).filter(StudentPoint.student_id.in_(student_ids)).all()
            for p in points:
                cong, nhac = points_map.get(p.student_id, (0, 0))
                if p.amount > 0:
                    cong += 1
                elif p.amount < 0:
                    nhac += 1
                points_map[p.student_id] = (cong, nhac)

    # Tạo excel
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    # Tính điểm phục vụ comment học tập
    student_scores: dict[int, dict[str, float]] = {sid: {} for sid in student_ids}
    student_overall_avg: dict[int, float | None] = {sid: None for sid in student_ids}

    for sid in student_ids:
        all_scores: list[float] = []
        subj_scores: dict[str, list[float]] = {}
        for subj_name, assigns in assignment_by_subject_name.items():
            for a in assigns:
                sub = sub_map.get((a.id, sid))
                score_val: float | None = None
                if sub is not None:
                    t = sub.teacher_score
                    ai = sub.ai_score
                    if t is not None and ai is not None:
                        score_val = (float(t) + float(ai)) / 2.0
                    elif t is None and ai is not None:
                        score_val = float(ai)
                    elif t is not None and ai is None:
                        score_val = float(t)
                if score_val is not None:
                    subj_scores.setdefault(subj_name, []).append(score_val)
                    all_scores.append(score_val)
        subj_avg = {k: (sum(v) / len(v)) for k, v in subj_scores.items() if v}
        student_scores[sid] = {k: float(v) for k, v in subj_avg.items()}
        if all_scores:
            student_overall_avg[sid] = sum(all_scores) / len(all_scores)

    # Sheet theo môn
    for subj_name, assigns in assignment_by_subject_name.items():
        ws = wb.create_sheet(title=_excel_safe_sheet_title(subj_name))
        ws.append(["Họ và tên"] + [f"Assignment {a.id}" for a in assigns])
        for s in students:
            sid = s.id
            row_vals: list = [s.full_name]
            for a in assigns:
                sub = sub_map.get((a.id, sid))
                score_val: float | None = None
                if sub is not None:
                    t = sub.teacher_score
                    ai = sub.ai_score
                    if t is not None and ai is not None:
                        score_val = (float(t) + float(ai)) / 2.0
                    elif t is None and ai is not None:
                        score_val = float(ai)
                    elif t is not None and ai is None:
                        score_val = float(t)
                row_vals.append(score_val if score_val is not None else "")
            ws.append(row_vals)

    # Sheet rèn luyện
    ws_rl = wb.create_sheet(title="Rèn Luyện")
    ws_rl.append(
        [
            "Họ và tên",
            "Số lần cộng điểm",
            "Số lần nhắc nhở",
            "Nhận xét về quá trình rèn luyện",
            "Nhận xét về quá trình học tập",
        ]
    )
    for s in students:
        cong, nhac = points_map.get(s.id, (0, 0))
        subj_avg_map = student_scores.get(s.id, {})
        overall_avg = student_overall_avg.get(s.id)

        # Fallback theo rule hiện tại
        best_subject = None
        if subj_avg_map:
            best_subject = max(subj_avg_map.items(), key=lambda kv: kv[1])[0]
        conduct_comment = _conduct_comment(int(cong), int(nhac))
        learning_comment = _learning_comment(best_subject, overall_avg)

        # Ưu tiên đánh giá bằng GPT
        try:
            conduct_comment, learning_comment = evaluate_conduct_and_learning_with_ai(
                student_name=s.full_name,
                positive_count=int(cong),
                reminder_count=int(nhac),
                subject_averages={k: float(v) for k, v in subj_avg_map.items()},
                overall_average=overall_avg,
            )
        except Exception:
            # Nếu AI không sẵn sàng/ lỗi parse/ timeout => fallback heuristics để không làm hỏng file.
            pass

        ws_rl.append(
            [
                s.full_name,
                cong,
                nhac,
                conduct_comment,
                learning_comment,
            ]
        )

    # Lưu & gửi
    export_dir = os.path.join("exports", "conduct")
    os.makedirs(export_dir, exist_ok=True)
    file_name = f"{safe_path_component(class_name)}_ren_luyen.xlsx"
    file_path = os.path.join(export_dir, file_name)
    wb.save(file_path)

    if q.message:
        with open(file_path, "rb") as f:
            await context.bot.send_document(
                chat_id=q.message.chat_id,
                document=f,
                filename=os.path.basename(file_path),
            )


async def handle_my_info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    q = update.callback_query
    if not q or not update.effective_user:
        return
    await q.answer()
    tg = update.effective_user
    with SessionLocal() as db:
        user = db.query(UserInfo).filter(UserInfo.telegram_id == str(tg.id)).first()
        if user is None:
            user = get_or_create_user(db, tg)
        else:
            db.refresh(user)
        # nạp quan hệ cho format thông tin
        db.expire(user, ["class_info", "homeroom_of_classes"])
        _ = user.class_info
        _ = user.homeroom_of_classes

    lines: list[str] = []
    lines.append(f"Họ tên: {user.full_name}")
    lines.append(f"tele_id: {user.telegram_id}")
    lines.append(f"username: {user.username or ''}")

    if user.role == UserRole.STUDENT:
        class_name = user.class_info.name if getattr(user, "class_info", None) else ""
        lines.insert(1, f"Lớp học: {class_name}")
        text = "👤 Thông tin học sinh\n\n" + "\n".join(lines)
    else:
        homerooms = getattr(user, "homeroom_of_classes", None) or []
        hr = ", ".join(c.name for c in homerooms) if homerooms else ""
        lines.append(f"Lớp đang chủ nhiệm: {hr}")
        text = "👤 Thông tin giáo viên\n\n" + "\n".join(lines)

    if q.message:
        await q.message.reply_text(text)


def register_handlers(app: Application) -> None:
    student_edit_name_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(student_edit_name_start, pattern=f"^{CB_S_EDIT}$")],
        states={
            WAITING_STUDENT_FULL_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, student_edit_name_receive),
            ],
        },
        fallbacks=[
            CommandHandler("cancel", student_edit_name_cancel),
            CommandHandler("start", student_edit_name_cancel),
            CommandHandler("menu", student_edit_name_cancel),
        ],
        name="student_edit_name",
        per_message=False,
    )

    teacher_edit_name_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(teacher_edit_name_start, pattern=f"^{CB_T_EDIT}$")],
        states={
            WAITING_TEACHER_FULL_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, teacher_edit_name_receive),
            ],
        },
        fallbacks=[
            CommandHandler("cancel", teacher_edit_name_cancel),
            CommandHandler("start", teacher_edit_name_cancel),
            CommandHandler("menu", teacher_edit_name_cancel),
        ],
        name="teacher_edit_name",
        per_message=False,
    )

    teacher_announce_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(teacher_announce_start, pattern=f"^{CB_T_ANN}$")],
        states={
            WAITING_TEACHER_ANNOUNCE_TEXT: [
                CallbackQueryHandler(teacher_announce_pick_class, pattern=r"^mt_ann_cls_\d+$"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, teacher_announce_receive_text),
            ],
        },
        fallbacks=[
            CommandHandler("cancel", teacher_announce_cancel),
            CommandHandler("start", teacher_announce_cancel),
            CommandHandler("menu", teacher_announce_cancel),
        ],
        name="teacher_announce",
        per_message=False,
    )

    # Callback cụ thể trước handler menu chung (cùng prefix ms_/mt_)
    app.add_handler(CallbackQueryHandler(on_student_pick_class, pattern=r"^ms_jc_\d+$"))
    app.add_handler(CallbackQueryHandler(on_teacher_join_resolve, pattern=r"^mt_ar_\d+_[ar]$"))
    app.add_handler(
        CallbackQueryHandler(
            on_teacher_pick_conduct_class,
            pattern=r"^mt_cond_cls_\d+$",
        )
    )

    app.add_handler(teacher_announce_conv)
    app.add_handler(teacher_edit_name_conv)
    app.add_handler(student_edit_name_conv)
    app.add_handler(CommandHandler("assign", teacher_assign_command))
    app.add_handler(CommandHandler("assignment", student_assignment_detail_command))
    app.add_handler(
        CommandHandler("viewAssigments", teacher_view_assignments_command)
    )
    app.add_handler(
        CommandHandler("detailAssignment", teacher_detail_assignment_command)
    )
    app.add_handler(
        CommandHandler(
            "viewSubmissionsAssginment",
            teacher_view_submissions_assignment_command,
        )
    )
    app.add_handler(
        CommandHandler("viewSubmisson", teacher_view_submission_command)
    )
    app.add_handler(CommandHandler("score", teacher_score_submission_command))
    app.add_handler(CommandHandler("plus", teacher_plus_command))
    app.add_handler(CommandHandler("minus", teacher_minus_command))
    app.add_handler(
        MessageHandler(
            filters.Document.PDF & filters.CaptionRegex(r"^/assign\b"),
            teacher_assign_command,
        )
    )
    app.add_handler(
        MessageHandler(
            filters.Document.PDF & ~filters.CaptionRegex(r"^/assign\b"),
            student_submit_assignment_pdf,
        )
    )
    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("menu", send_role_menu))
    app.add_handler(CallbackQueryHandler(on_menu_button, pattern=r"^(mt_|ms_)"))
